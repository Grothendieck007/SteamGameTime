import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from steam import Steam
from dotenv import load_dotenv
import os

def main_function():
    # 加载环境变量
    load_dotenv(dotenv_path="mysteamapi.env")

    # 创建Steam API实例
    KEY = os.getenv("STEAM_API_KEY")
    steam = Steam(KEY)

    # 从配置文件中读取Steam ID
    with open("config.txt", "r") as file:
        steamid = file.read().strip()

    try:
        # 获取用户拥有的游戏信息
        response = steam.users.get_owned_games(steamid)
        owned_games = response["games"]

        filtered_games = []

        # 遍历每个游戏
        for game in owned_games:
            appid = game["appid"]
            playtime_forever = game["playtime_forever"]

            if playtime_forever == 0:
                continue  # 跳过playtime_forever为0的游戏

            name = game.get("name", "未知游戏")

            game_info = {
                "name": name,
                "playtime_forever": playtime_forever
            }

            filtered_games.append(game_info)

        # 将筛选过的游戏信息存储为game.json文件
        data = {
            "用户拥有的游戏": filtered_games
        }

        with open("game.json", "w", encoding="utf-8") as file:
            json.dump(data, file, ensure_ascii=False)
            print("数据已成功存储到game.json文件。")

        # 从JSON文件中读取数据
        with open("game.json", "r", encoding="utf-8") as file:
            data = json.load(file)

        # 提取用户拥有的游戏信息
        game_data = data["用户拥有的游戏"]

        # 创建一个新的Excel工作簿
        workbook = Workbook()

        # 获取默认的活动工作表
        sheet = workbook.active

        # 设置A列的宽度为53
        sheet.column_dimensions[get_column_letter(1)].width = 53

        # 设置B列的宽度为18
        sheet.column_dimensions[get_column_letter(2)].width = 18

        sheet["A1"] = "游戏名称"
        sheet["B1"] = "游戏时间（小时）"

        # 逐行添加数据
        row = 2
        for game in game_data:
            name = game["name"]
            playtime_minutes = game["playtime_forever"]
            playtime_hours = round(playtime_minutes / 60, 1)  # 将小时保留一位小数
            sheet[f"A{row}"] = name
            sheet[f"B{row}"] = f"{playtime_hours}小时"
            row += 1

        # 设置A1内容居中对齐
        sheet["A1"].font = Font(bold=True)
        sheet["A1"].alignment = Alignment(horizontal="center")

        # 设置B列内容居中对齐
        for row in sheet.iter_cols(min_col=2, max_col=2):
            for cell in row:
                cell.alignment = Alignment(horizontal="center")

        # 将数据保存为Excel文件
        workbook.save("game_data.xlsx")
        print("数据已成功保存为game_data.xlsx文件。")

    except Exception as e:
        print("代码中止运行。")
        raise e